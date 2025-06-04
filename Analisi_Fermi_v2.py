import pandas as pd  # pandas per la manipolazione dei dati
from openpyxl import load_workbook  # per modificare i fogli excel esistenti
import matplotlib.pyplot as plt
import seaborn as sns #per disegnare grafici statistici

# Percorso file Excel
#percorso_file = r'C:\Users\bsanzi\OneDrive - FIAMM Energy Technology S.p.A\Desktop\Trim_01_25_Linee123.xlsx'
percorso_file = r'C:\Users\Utente\Desktop\Trim_01_25_Linee123.xls.xlsx'
linea_da_analizzare = input ('Quale linea analizzo?')
macchina_rul = input ('Per quale macchina desideri calcolare variabilità della RUL? ')

# Nomi colonne
nome_colonna_data = 'DATA'
nome_colonna_linea = 'LINEA'
nome_colonna_macchina = 'MACCHINA'
nome_colonna_pezzo = 'PEZZO'
nome_colonna_stato_pezzo = 'STATO PEZZO'
nome_colonna_classe = 'CLASSE'
nome_colonna_minuti = 'MINUTI'

# Nomi fogli output da scrivere
nome_foglio_scrittura_fermi_accorpati = 'fermi_accorpati_per_macchina'
nome_foglio_scrittura_analisi_macchina = 'analisi_fermi_per_macchina'
nome_foglio_scrittura_analisi_pezzo = 'analisi_fermi_per_pezzo'
nome_foglio_scrittura_analisi_classe = 'analisi_fermi_per_classe'
#nome_foglio_scrittura_analisi_rul = '' è ora dinamico, quindi non è più una costante qui

# Nome foglio input
nome_foglio_lettura = 'Estraz_Ev'

try:
    # Leggi il foglio Excel
    df = pd.read_excel(percorso_file, sheet_name=nome_foglio_lettura)

    # Filtra solo i dati della linea richiesta
    df = df[df[nome_colonna_linea].astype(str).str.strip() == linea_da_analizzare.strip()]

    # Verifica se ci sono dati
    if df.empty:
        raise ValueError(f"Nessun dato trovato per la linea '{linea_da_analizzare}'. Interrompo l'elaborazione.")

    # Verifica colonne richieste
    colonne_richieste = [nome_colonna_data, nome_colonna_linea, nome_colonna_macchina, nome_colonna_pezzo, nome_colonna_stato_pezzo, nome_colonna_classe, nome_colonna_minuti]

    for col in colonne_richieste:
        if col not in df.columns:
            raise KeyError(f"La colonna '{col}' non è stata trovata nel foglio '{nome_foglio_lettura}'.")
        
    # stato processamento programma
    print("loading 25%")


    #accorpamento per cambio produzione 
    df_cambio = df[df[nome_colonna_macchina].str.lower().str.strip() == "cambio produzione"].copy()

    if not df_cambio.empty:
        df_cambio[nome_colonna_data] = pd.to_datetime(df_cambio[nome_colonna_data], errors='coerce')
        df_cambio['DATA ACC'] = df_cambio[nome_colonna_data].dt.date

    # Somma minuti per giorno
        minuti_per_giorno = df_cambio.groupby('DATA ACC')[nome_colonna_minuti].sum().sort_index()

        accorpati = []
        giorni_processati = set()

        for giorno in minuti_per_giorno.index:
            if giorno in giorni_processati:
                continue

            giorno_successivo = giorno + pd.Timedelta(days=1)
            minuti_oggi = minuti_per_giorno[giorno]
            minuti_domani = minuti_per_giorno.get(giorno_successivo, 0)

            accorpati.append({
                'DATA_CAMBIO': giorno,
                'MINUTI_TOT_CUMULATI': minuti_oggi + minuti_domani
            })

            giorni_processati.add(giorno)
            giorni_processati.add(giorno_successivo)

        df_cambi_produzione = pd.DataFrame(accorpati)
        df_cambi_produzione.sort_values(by='DATA_CAMBIO', inplace=True)

        nome_foglio_cambi_produzione = f'cambi_prod_L{linea_da_analizzare}'


    # Pulizia dei dati
    df = df.dropna(subset=[nome_colonna_minuti, nome_colonna_macchina, nome_colonna_pezzo]) # rimuove le righe con valori mancanti
    df[nome_colonna_data] = pd.to_datetime(df[nome_colonna_data]) # Converte la colonna DATA FERMO in tipo datetime
    df[nome_colonna_minuti] = pd.to_numeric(df[nome_colonna_minuti], errors='coerce') #Conversione Tipo Numerico, gli errori diventano NaN
    df = df.dropna(subset=[nome_colonna_minuti]) # Rimuovi righe dove 'MINUTI' è diventato NaN dopo la conversione


    # Accorpamento per giorno
    df['DATA ACC'] = df[nome_colonna_data].dt.date # crea una nuova colonna DATA ACC contenente solo la data senza l'ora
    idx_max = df.groupby(['DATA ACC', nome_colonna_macchina])[nome_colonna_minuti].idxmax() # trova l'indice della riga con i minuti di fermo maggiori
    fermi_accorpati = df.loc[idx_max].copy() # estrae le righe con il fermo più lungo. copy-per creare un nuovo dataframe

    # Somma dei minuti per (giorno, macchina)
    somma_minuti = df.groupby(['DATA ACC', nome_colonna_macchina])[nome_colonna_minuti].sum()
    fermi_accorpati[nome_colonna_minuti] = fermi_accorpati.set_index(['DATA ACC', nome_colonna_macchina]).index.map(somma_minuti)

    # Ordina e pulizia
    fermi_accorpati = fermi_accorpati.sort_values(by=[nome_colonna_data, nome_colonna_macchina])
    # Elimina la colonna ausiliaria una sola volta
    fermi_accorpati.drop(columns=['DATA ACC'], inplace=True)
    # Converti la colonna 'Data turno' al solo formato data
    fermi_accorpati[nome_colonna_data] = pd.to_datetime(fermi_accorpati[nome_colonna_data], errors='coerce')
    
    nome_foglio_scrittura_fermi_accorpati = f'fermi_linea_{linea_da_analizzare}'


    # === Periodo analizzato (basato su dati accorpati) ===
    totale_fermi = len(fermi_accorpati)
    data_inizio = fermi_accorpati[nome_colonna_data].min()
    data_fine = fermi_accorpati[nome_colonna_data].max()
    durata_periodo_giorni = (data_fine - data_inizio).days + 1

     # === Analisi per macchina ===
     # df_macchina_rul_data = fermi_accorpati[fermi_accorpati[nome_colonna_macchina] == macchina_rul].copy()
    analisi_fermi_macchina = fermi_accorpati.groupby(nome_colonna_macchina).agg(
        numero_totale_fermi=(nome_colonna_pezzo, 'count'),
        tempo_totale_fermo=(nome_colonna_minuti, 'sum')
    ).reset_index()
    analisi_fermi_macchina['frequenza_guasto (g)'] = analisi_fermi_macchina['numero_totale_fermi'] / durata_periodo_giorni
    analisi_fermi_macchina['MTBF (g)'] = 1 / analisi_fermi_macchina['frequenza_guasto (g)']
    analisi_fermi_macchina = analisi_fermi_macchina.sort_values(by='tempo_totale_fermo', ascending=False)

    # === Analisi per pezzo ===
    analisi_fermi_pezzo = fermi_accorpati.groupby([nome_colonna_macchina, nome_colonna_pezzo]).agg(
        numero_totale_fermi=(nome_colonna_minuti, 'count'),
        tempo_totale_fermo=(nome_colonna_minuti, 'sum')
    ).reset_index()
    analisi_fermi_pezzo['frequenza_guasto (g)'] = analisi_fermi_pezzo['numero_totale_fermi'] / durata_periodo_giorni
    analisi_fermi_pezzo['MTBF (g)'] = 1 / analisi_fermi_pezzo['frequenza_guasto (g)']
    analisi_fermi_pezzo = analisi_fermi_pezzo.sort_values(by='tempo_totale_fermo', ascending=False)

    # stato processamento programma
    print("loading 50%")

    # === Analisi per classe ===
    analisi_fermi_classe = fermi_accorpati.groupby([nome_colonna_macchina,nome_colonna_pezzo, nome_colonna_classe]).agg(
        numero_totale_fermi=(nome_colonna_minuti, 'count'),
        tempo_totale_fermo=(nome_colonna_minuti, 'sum')
    ).reset_index()
    analisi_fermi_classe['frequenza_guasto (g)'] = analisi_fermi_classe['numero_totale_fermi'] / durata_periodo_giorni
    analisi_fermi_classe['MTBF (g)'] = 1 / analisi_fermi_classe['frequenza_guasto (g)']
    analisi_fermi_classe = analisi_fermi_classe.sort_values(by='tempo_totale_fermo', ascending=False)


    # === Calcolo e creazione foglio RUL per la macchina specificata da macchina_rul ===

    df_macchina_rul_data = fermi_accorpati[fermi_accorpati[nome_colonna_macchina] == macchina_rul].copy()

    if not df_macchina_rul_data.empty:
        # Ordina per data
        df_macchina_rul_data = df_macchina_rul_data.sort_values(by=nome_colonna_data)

        # Calcola differenza giorni
        df_macchina_rul_data['rul'] = df_macchina_rul_data[nome_colonna_data].diff().dt.total_seconds() / (60 * 60 * 24)
        df_macchina_rul_data['rul'] = df_macchina_rul_data['rul'].fillna(0).round(2)
        
        # Costruisci un DataFrame ridotto
        df_rul_compatta = df_macchina_rul_data[[nome_colonna_data, 'rul', nome_colonna_minuti]].copy()
        df_rul_compatta.columns = ['Data', 'rul', 'minuti di fermo']

        # Aggiungi una riga vuota sopra con il nome della macchina
        macchina_header_rul = pd.DataFrame({'Data': [f'Macchina: {macchina_rul}'], 'Data': [''], 'rul': [None], 'minuti di fermo': [None]})

        df_finale_rul = pd.concat([macchina_header_rul, df_rul_compatta], ignore_index=True)
        nome_foglio_scrittura_analisi_rul_specifica = f"RUL_{macchina_rul}"

    else:
        print(f"Nessun dato di fermo trovato per la macchina '{macchina_rul}' per calcolare il rul.")

################# PARTE NUOVA ################
# Unione delle date di Cambio Produzione e RUL
    df_unione_date = pd.DataFrame() # Inizializza un DataFrame vuoto per le date unite

    # Assicurati che df_cambi_produzione esista e sia valido
    if 'df_cambi_produzione' in locals() and not df_cambi_produzione.empty:
        
    # Prepara df_cambi_produzione per l'unione: converti la colonna DATA_CAMBIO a solo data
        df_cambi_produzione['DATA_CAMBIO'] = pd.to_datetime(df_cambi_produzione['DATA_CAMBIO']).dt.date
    # Rinomina la colonna dei minuti per distinguerla dopo l'unione
        df_cambi_produzione_per_merge = df_cambi_produzione[['DATA_CAMBIO', 'MINUTI_TOT_CUMULATI']].copy()
        df_cambi_produzione_per_merge.rename(columns={'DATA_CAMBIO': 'Data', 'MINUTI_TOT_CUMULATI': 'Minuti_Cambio_Produzione'}, inplace=True)
        df_unione_date = df_cambi_produzione_per_merge # La prima base per l'unione

    # Assicurati che df_rul_compatta esista e sia valido
    if 'df_rul_compatta' in locals() and not df_rul_compatta.empty:

        # Uniforma il tipo di dato della colonna 'Data'
        df_rul_compatta['Data'] = pd.to_datetime(df_rul_compatta['Data']).dt.date

        if df_unione_date.empty:
            # Se df_cambi_produzione non esisteva, inizia con le date RUL
           df_unione_date = df_rul_compatta.copy()
        else:
        # Unisci i DataFrame sulle colonne della data
        # 'outer' join per includere tutte le date da entrambi i lati
            df_unione_date = pd.merge(df_unione_date, df_rul_compatta, on='Data', how='outer')

# Ordina il DataFrame unito per data e pulisci eventuali righe vuote create dal merge iniziale
    if not df_unione_date.empty:
        df_unione_date = df_unione_date.sort_values(by='Data').reset_index(drop=True)
        # Rimuovi righe con valori NaN nella colonna 'Data' se presenti
        df_unione_date.dropna(subset=['Data'], inplace=True)

        # Riempie i valori NaN con 0
        df_unione_date.fillna(0, inplace=True)
  

    # Nome del foglio di output per i dati uniti
    nome_foglio_scrittura_date_unite = f'CAMBIO PRODUZIONE_RUL_L{linea_da_analizzare}_{macchina_rul}'


    # === Rimozione fogli esistenti prima di scrivere i nuovi ===
    workbook = load_workbook(percorso_file)
    fogli_da_rimuovere = [
        nome_foglio_scrittura_fermi_accorpati,
        nome_foglio_scrittura_analisi_macchina,
        nome_foglio_scrittura_analisi_pezzo,
        nome_foglio_scrittura_analisi_classe
    
    ]
    if 'df_finale_rul' in locals(): # Rimuovi il foglio RUL specifico solo se è stato creato
        if nome_foglio_scrittura_analisi_rul_specifica in workbook.sheetnames:
            std = workbook[nome_foglio_scrittura_analisi_rul_specifica]
            workbook.remove(std)
    # stato processamento programma
    print("loading 75%")

    if 'nome_foglio_scrittura_date_unite' in locals():
         if nome_foglio_scrittura_date_unite in workbook.sheetnames:
             std = workbook[nome_foglio_scrittura_date_unite]
             workbook.remove(std)

    for nome_foglio in fogli_da_rimuovere:
        if nome_foglio in workbook.sheetnames:
            std = workbook[nome_foglio]
            workbook.remove(std)
            workbook.save(percorso_file)

    # === Scrittura su Excel ===
    with pd.ExcelWriter(percorso_file, engine='openpyxl', mode='a') as writer:
        df_cambi_produzione.to_excel(writer, sheet_name=nome_foglio_cambi_produzione, index=False)
        fermi_accorpati.to_excel(writer, sheet_name=nome_foglio_scrittura_fermi_accorpati, index=False)
        analisi_fermi_macchina.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_macchina, index=False)
        analisi_fermi_pezzo.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_pezzo, index=False)
        analisi_fermi_classe.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_classe, index=False)
        if 'df_finale_rul' in locals(): # Scrivi il foglio RUL solo se è stato creato
            df_finale_rul.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_rul_specifica, index=False)
        if 'df_unione_date' in locals() and not df_unione_date.empty:
            df_unione_date.to_excel(writer, sheet_name=nome_foglio_scrittura_date_unite, index=False)

    
    # === Calcolo e scrittura variabilità della RUL nel foglio Excel ===
    if 'df_finale_rul' in locals():
        variabilita_rul = df_macchina_rul_data['rul'].std().round(2) #se esiste la rul, calcola la deviazione standard e formatta il risultato con 2 cifre decimali

        workbook = load_workbook(percorso_file)
        sheet_rul = workbook[nome_foglio_scrittura_analisi_rul_specifica]

        # Scrivi nella riga 1 alla prima colonna libera
        prima_riga = sheet_rul[1]
        prima_colonna_libera = len(prima_riga) + 2  # Conta colonne usate e aggiunge 2

        col_label = sheet_rul.cell(row=1, column=prima_colonna_libera)
        col_value = sheet_rul.cell(row=2, column=prima_colonna_libera)  # Riga 2 perché riga 1 è intestazione

        col_label.value = "Variabilità RUL (std):"
        col_value.value = variabilita_rul

        workbook.save(percorso_file)

    if not df_unione_date.empty:
        # Prepara il DataFrame per il grafico
        df_plot = df_unione_date.copy()
        df_plot['Data'] = pd.to_datetime(df_plot['Data'])  # Converte Data per assi temporali

        # Crea la figura con due assi Y
        fig, ax1 = plt.subplots(figsize=(14, 6))

        # Asse Y sinistro: RUL (linea blu)
        color_line = 'tab:blue'
        ax1.set_xlabel('Data')
        ax1.set_ylabel('RUL (giorni)', color=color_line)
        ax1.plot(df_plot['Data'], df_plot['rul'], color=color_line, label='RUL', linewidth=2)
        ax1.tick_params(axis='y', labelcolor=color_line)

        # Asse Y destro: Minuti (colonne rosse e verdi)
        ax2 = ax1.twinx()
        ax2.set_ylabel('minuti di fermo', color='gray')
        ax2.bar(df_plot['Data'], df_plot['minuti di fermo'], color='red', label='Minuti Fermo', alpha=0.6)
        ax2.bar(df_plot['Data'], df_plot['Minuti_Cambio_Produzione'], color='green', label='Minuti Cambio Produzione', alpha=0.4)
        ax2.tick_params(axis='y', labelcolor='gray')

        # Migliora layout e legende
        fig.autofmt_xdate()
        fig.tight_layout()
        plt.title(f"RUL_Minuti di Fermo_Cambio Produzione - Linea {linea_da_analizzare}, Macchina {macchina_rul}")

        lines_labels = ax1.get_legend_handles_labels()
        bars_labels = ax2.get_legend_handles_labels()
        ax1.legend(lines_labels[0] + bars_labels[0], lines_labels[1] + bars_labels[1], loc='upper left')

        # Salva il grafico in un file PNG accanto all'Excel
        grafico_path = percorso_file.replace('.xlsx', f'_{macchina_rul}_grafico.png')
        plt.savefig(grafico_path)
        plt.show()
        plt.close()

    print(f"Grafico combinato salvato in: {grafico_path}")


    # === Output a terminale ===
    print(f"\n--- Riepilogo Analisi Fermo Macchine ---")
    print(f"Periodo analizzato: dal **{data_inizio.date()}** al **{data_fine.date()}** (**{durata_periodo_giorni} giorni**)")
    print(f"Totale fermi accorpati (macchina/giorno): **{totale_fermi}**")
    print(f"Dati accorpati per macchina/giorno salvati sul foglio: '**{nome_foglio_scrittura_fermi_accorpati}**'")
    print(f"Analisi per macchina salvata sul foglio: '**{nome_foglio_scrittura_analisi_macchina}**'")
    print(f"Analisi per pezzo salvata sul foglio: '**{nome_foglio_scrittura_analisi_pezzo}**'")
    if 'df_finale_rul' in locals():
        print(f"Analisi RUL per la macchina '{macchina_rul}' salvata sul foglio: '**{nome_foglio_scrittura_analisi_rul_specifica}**'")
    print(f"\nElaborazione completata con successo. Controlla il file: {percorso_file}")
    print(f"Variabilità (dev. standard) della RUL per la macchina '{macchina_rul}': {variabilita_rul} giorni")
    

except FileNotFoundError:
    print(f"Errore: Il file non è stato trovato al percorso specificato: '{percorso_file}'")
except KeyError as e:
    print(f"Errore nella lettura delle colonne: {e}")
except Exception as e:
    print(f"Si è verificato un errore inatteso: {e}")