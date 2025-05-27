import pandas as pd  # pandas per la manipolazione dei dati
from openpyxl import load_workbook  # per modificare i fogli excel esistenti
import matplotlib.pyplot as plt
#import seaborn as sns #per disegnare grafici statistici

# Percorso file Excel
percorso_file = r'C:\Users\bsanzi\OneDrive - FIAMM Energy Technology S.p.A\Desktop\fermitot.xlsx'
linea_da_analizzare = input ('Quale linea analizzo?')
macchina_rul = input ('Per quale macchina desideri calcolare il RUL? ')

# Nomi colonne
nome_colonna_data_fermo = 'DATA FERMO'
nome_colonna_data_turno = 'DATA TURNO'
nome_colonna_linea = 'LINEA/FONDIT.'
nome_colonna_macchina = 'MACCHINA'
nome_colonna_pezzo = 'PEZZO'
nome_colonna_stato_pezzo = 'STATO PEZZO'
nome_colonna_classe = 'CLASSE'
nome_colonna_minuti = 'MINUTI'

# Nomi fogli output da scrivere
nome_foglio_scrittura_fermi_accorpati = 'fermi_accorpati_per_macchina'
nome_foglio_scrittura_analisi_macchina = 'analisi_fermi_per_macchina'
nome_foglio_scrittura_analisi_pezzo = 'analisi_fermi_per_pezzo'
#nome_foglio_scrittura_analisi_rul = '' è ora dinamico, quindi non è più una costante qui

# Nome foglio input
nome_foglio_lettura = 'Estraz_Ev'

try:
    # Leggi il foglio Excel
    df = pd.read_excel(percorso_file, sheet_name=nome_foglio_lettura)

    # Filtra solo i dati della linea richiesta
    df = df[df[nome_colonna_linea].astype(str).str.strip() == linea_da_analizzare.strip()]

    # Verifica se ci sono dati
    if df.empty:
        raise ValueError(f"Nessun dato trovato per la linea '{linea_da_analizzare}'. Interrompo l'elaborazione.")

    # Verifica colonne richieste
    colonne_richieste = [nome_colonna_data_fermo, nome_colonna_data_turno, nome_colonna_linea, nome_colonna_macchina, nome_colonna_pezzo, nome_colonna_stato_pezzo, nome_colonna_classe, nome_colonna_minuti]

    for col in colonne_richieste:
        if col not in df.columns:
            raise KeyError(f"La colonna '{col}' non è stata trovata nel foglio '{nome_foglio_lettura}'.")
        
    #1111111111111111111111111111111 
    print("dati input trovati")

    # Pulizia dei dati
    df = df.dropna(subset=[nome_colonna_minuti, nome_colonna_macchina, nome_colonna_pezzo]) # rimuove le righe con valori mancanti
    df[nome_colonna_data_fermo] = pd.to_datetime(df[nome_colonna_data_fermo]) # Converte la colonna DATA FERMO in tipo datetime
    df[nome_colonna_data_turno] = pd.to_datetime(df[nome_colonna_data_turno]).dt.date
    df[nome_colonna_minuti] = pd.to_numeric(df[nome_colonna_minuti], errors='coerce') #Conversione Tipo Numerico, gli errori diventano NaN
    df = df.dropna(subset=[nome_colonna_minuti]) # Rimuovi righe dove 'MINUTI' è diventato NaN dopo la conversione

    # Accorpamento per giorno
    df['DATA ACC'] = df[nome_colonna_data_fermo].dt.date # crea una nuova colonna DATA ACC contenente solo la data senza l'ora
    idx_max = df.groupby(['DATA ACC', nome_colonna_macchina])[nome_colonna_minuti].idxmax() # trova l'indice della riga con i minuti di fermo maggiori
    fermi_accorpati = df.loc[idx_max].copy() # estrae le righe con il fermo più lungo. copy-per creare un nuovo dataframe

    # Somma dei minuti per (giorno, macchina)
    somma_minuti = df.groupby(['DATA ACC', nome_colonna_macchina])[nome_colonna_minuti].sum()
    fermi_accorpati[nome_colonna_minuti] = fermi_accorpati.set_index(['DATA ACC', nome_colonna_macchina]).index.map(somma_minuti)

    # Ordina e pulizia
    fermi_accorpati = fermi_accorpati.sort_values(by=[nome_colonna_data_fermo, nome_colonna_macchina])
    # Elimina la colonna ausiliaria una sola volta
    fermi_accorpati.drop(columns=['DATA ACC'], inplace=True)
    # Converti la colonna 'Data turno' al solo formato data
    fermi_accorpati[nome_colonna_data_turno] = pd.to_datetime(fermi_accorpati[nome_colonna_data_turno], errors='coerce').dt.date
    
    nome_foglio_scrittura_fermi_accorpati = f'fermi_linea_{linea_da_analizzare}'

    # === Periodo analizzato (basato su dati accorpati) ===
    totale_fermi = len(fermi_accorpati)
    data_inizio = fermi_accorpati[nome_colonna_data_fermo].min()
    data_fine = fermi_accorpati[nome_colonna_data_fermo].max()
    durata_periodo_giorni = (data_fine - data_inizio).days + 1

     # === Analisi per macchina ===
    analisi_fermi_macchina = fermi_accorpati.groupby(nome_colonna_macchina).agg(
        numero_totale_fermi=(nome_colonna_pezzo, 'count'),
        tempo_totale_fermo=(nome_colonna_minuti, 'sum')
    ).reset_index()
    analisi_fermi_macchina['frequenza_guasto (g)'] = analisi_fermi_macchina['numero_totale_fermi'] / durata_periodo_giorni
    analisi_fermi_macchina['MTBF'] = 1 / analisi_fermi_macchina['frequenza_guasto (g)']
    analisi_fermi_macchina = analisi_fermi_macchina.sort_values(by='tempo_totale_fermo', ascending=False)

    # === Analisi per pezzo ===
    analisi_fermi_pezzo = fermi_accorpati.groupby([nome_colonna_macchina, nome_colonna_pezzo]).agg(
        numero_totale_fermi=(nome_colonna_minuti, 'count'),
        tempo_totale_fermo=(nome_colonna_minuti, 'sum')
    ).reset_index()
    analisi_fermi_pezzo['frequenza_guasto (g)'] = analisi_fermi_pezzo['numero_totale_fermi'] / durata_periodo_giorni
    analisi_fermi_pezzo['MTBF'] = 1 / analisi_fermi_pezzo['frequenza_guasto (g)']
    analisi_fermi_pezzo = analisi_fermi_pezzo.sort_values(by='tempo_totale_fermo', ascending=False)

    # === Calcolo e creazione foglio RUL per la macchina specificata da macchina_rul ===
    df_macchina_rul_data = fermi_accorpati[fermi_accorpati[nome_colonna_macchina] == macchina_rul].copy()

    if not df_macchina_rul_data.empty:
        # Ordina per data
        df_macchina_rul_data = df_macchina_rul_data.sort_values(by=nome_colonna_data_fermo)

        # Calcola differenza giorni
        df_macchina_rul_data['rul'] = df_macchina_rul_data[nome_colonna_data_fermo].diff().dt.total_seconds() / (60 * 60 * 24)
        df_macchina_rul_data['rul'] = df_macchina_rul_data['rul'].fillna(0).round(2)
        
        # Costruisci un DataFrame ridotto
        df_rul_compatta = df_macchina_rul_data[[nome_colonna_data_fermo, nome_colonna_data_turno, 'rul', nome_colonna_minuti]].copy()
        df_rul_compatta.columns = ['Data fermi','Data turno', 'rul', 'minuti']

        # Aggiungi una riga vuota sopra con il nome della macchina
        macchina_header_rul = pd.DataFrame({'Data fermi': [f'Macchina: {macchina_rul}'], 'Data turno': [''], 'rul': [None], 'minuti': [None]})

        df_finale_rul = pd.concat([macchina_header_rul, df_rul_compatta], ignore_index=True)
        nome_foglio_scrittura_analisi_rul_specifica = f"RUL_{macchina_rul}"

    else:
        print(f"Nessun dato di fermo trovato per la macchina '{macchina_rul}' per calcolare il rul.")


    # === Rimozione fogli esistenti prima di scrivere i nuovi ===
    workbook = load_workbook(percorso_file)
    fogli_da_rimuovere = [
        nome_foglio_scrittura_fermi_accorpati,
        nome_foglio_scrittura_analisi_macchina,
        nome_foglio_scrittura_analisi_pezzo
    
    ]
    if 'df_finale_rul' in locals(): # Rimuovi il foglio RUL specifico solo se è stato creato
        if nome_foglio_scrittura_analisi_rul_specifica in workbook.sheetnames:
            std = workbook[nome_foglio_scrittura_analisi_rul_specifica]
            workbook.remove(std)
    
    for nome_foglio in fogli_da_rimuovere:
        if nome_foglio in workbook.sheetnames:
            std = workbook[nome_foglio]
            workbook.remove(std)
    workbook.save(percorso_file)

    # === Scrittura su Excel ===
    with pd.ExcelWriter(percorso_file, engine='openpyxl', mode='a') as writer:
        fermi_accorpati.to_excel(writer, sheet_name=nome_foglio_scrittura_fermi_accorpati, index=False)
        analisi_fermi_macchina.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_macchina, index=False)
        analisi_fermi_pezzo.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_pezzo, index=False)
        if 'df_finale_rul' in locals(): # Scrivi il foglio RUL solo se è stato creato
            df_finale_rul.to_excel(writer, sheet_name=nome_foglio_scrittura_analisi_rul_specifica, index=False)

    
    # === Calcolo e scrittura variabilità della RUL nel foglio Excel ===
    if 'df_finale_rul' in locals():
        variabilita_rul = df_macchina_rul_data['rul'].std().round(2) #se esiste la rul, calcola la deviazione standard e formatta il risultato con 2 cifre decimali

        workbook = load_workbook(percorso_file)
        sheet_rul = workbook[nome_foglio_scrittura_analisi_rul_specifica]

        # Scrivi nella riga 1 alla prima colonna libera
        prima_riga = sheet_rul[1]
        prima_colonna_libera = len(prima_riga) + 2  # Conta colonne usate e aggiunge 2

        col_label = sheet_rul.cell(row=1, column=prima_colonna_libera)
        col_value = sheet_rul.cell(row=2, column=prima_colonna_libera)  # Riga 2 perché riga 1 è intestazione

        col_label.value = "Variabilità RUL (std):"
        col_value.value = variabilita_rul

        workbook.save(percorso_file)


    # === Output a terminale ===
    print(f"\n--- Riepilogo Analisi Fermo Macchine ---")
    print(f"Periodo analizzato: dal **{data_inizio.date()}** al **{data_fine.date()}** (**{durata_periodo_giorni} giorni**)")
    print(f"Totale fermi accorpati (macchina/giorno): **{totale_fermi}**")
    print(f"Dati accorpati per macchina/giorno salvati sul foglio: '**{nome_foglio_scrittura_fermi_accorpati}**'")
    print(f"Analisi per macchina salvata sul foglio: '**{nome_foglio_scrittura_analisi_macchina}**'")
    print(f"Analisi per pezzo salvata sul foglio: '**{nome_foglio_scrittura_analisi_pezzo}**'")
    if 'df_finale_rul' in locals():
        print(f"Analisi RUL per la macchina '{macchina_rul}' salvata sul foglio: '**{nome_foglio_scrittura_analisi_rul_specifica}**'")
    print(f"\nElaborazione completata con successo. Controlla il file: {percorso_file}")
    print(f"Variabilità (dev. standard) della RUL per la macchina '{macchina_rul}': {variabilita_rul} giorni")
    

except FileNotFoundError:
    print(f"Errore: Il file non è stato trovato al percorso specificato: '{percorso_file}'")
except KeyError as e:
    print(f"Errore nella lettura delle colonne: {e}")
except Exception as e:
    print(f"Si è verificato un errore inatteso: {e}")
